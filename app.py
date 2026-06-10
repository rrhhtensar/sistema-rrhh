from flask import Flask, render_template, request, jsonify, send_file, session
import pandas as pd
import json
import io
import os
import unicodedata
import re
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'rrhh-tensar-2026')

# ═══════════════════════════════════════════════
#  FERIADOS ARGENTINA
# ═══════════════════════════════════════════════
FERIADOS = {
    date(2026,1,1), date(2026,2,16), date(2026,2,17),
    date(2026,3,24), date(2026,4,2), date(2026,4,3), date(2026,4,4),
    date(2026,5,1), date(2026,5,25),
    date(2026,6,15), date(2026,6,20),
    date(2026,7,9), date(2026,8,17),
    date(2026,10,12), date(2026,11,20),
    date(2026,12,8), date(2026,12,25),
    date(2025,1,1), date(2025,3,3), date(2025,3,4),
    date(2025,3,24), date(2025,4,2), date(2025,4,18), date(2025,4,19),
    date(2025,5,1), date(2025,5,25), date(2025,6,16), date(2025,6,20),
    date(2025,7,9), date(2025,8,18), date(2025,10,12),
    date(2025,11,17), date(2025,11,24),
    date(2025,12,8), date(2025,12,25),
}

NOMBRES_MES = ['','Enero','Febrero','Marzo','Abril','Mayo','Junio',
               'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

# ═══════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════
def normalizar(texto):
    texto = str(texto).upper().strip()
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
    return re.sub(r'[^A-Z0-9 ]', '', texto).strip()

def similitud_nombre(n1, n2):
    t1 = set(normalizar(n1).split())
    t2 = set(normalizar(n2).split())
    if not t1 or not t2:
        return 0
    return len(t1 & t2) / max(len(t1), len(t2))

def parse_hora(h_str):
    h_str = str(h_str).strip()
    if 'Sin reg' in h_str or h_str in ('', 'nan', 'None'):
        return None
    try:
        partes = h_str.split(':')
        return int(partes[0]) * 60 + int(partes[1])
    except:
        return None

def mins_a_hs(m):
    return round(m / 60, 2) if m and m > 0 else 0.0

def clasificar_dia(fecha):
    if fecha in FERIADOS:
        return 'feriado'
    dow = fecha.weekday()
    if dow == 6: return 'domingo'
    if dow == 5: return 'sabado'
    if dow == 4: return 'viernes'
    return 'normal'

def calcular_horas_dia(entrada_min, salida_min, tipo_dia, jornada_12=False):
    res = dict(hs_normales=0.0, hs_ext50=0.0, hs_ext100=0.0,
               marcacion_incompleta=False, ausente=False)
    if entrada_min is None and salida_min is None:
        res['ausente'] = True
        return res
    if entrada_min is None or salida_min is None:
        res['marcacion_incompleta'] = True
        return res
    trabajado = salida_min - entrada_min
    if trabajado <= 0:
        return res

    if tipo_dia in ('feriado', 'domingo'):
        res['hs_ext100'] = mins_a_hs(trabajado)
    elif tipo_dia == 'sabado':
        limite_13 = 13 * 60
        if salida_min <= limite_13:
            res['hs_ext50'] = mins_a_hs(trabajado)
        elif entrada_min >= limite_13:
            res['hs_ext100'] = mins_a_hs(trabajado)
        else:
            res['hs_ext50']  = mins_a_hs(limite_13 - entrada_min)
            res['hs_ext100'] = mins_a_hs(salida_min - limite_13)
    elif tipo_dia == 'viernes':
        if jornada_12:
            normales_tope = 8 * 60
            if trabajado <= normales_tope:
                res['hs_normales'] = mins_a_hs(trabajado)
            else:
                res['hs_normales'] = mins_a_hs(normales_tope)
                res['hs_ext50']    = mins_a_hs(trabajado - normales_tope)
        else:
            normales_tope = 8 * 60
            base           = 8 * 60 + 30
            if trabajado <= normales_tope:
                res['hs_normales'] = mins_a_hs(trabajado)
            elif trabajado <= base:
                res['hs_normales'] = mins_a_hs(normales_tope)
            else:
                res['hs_normales'] = mins_a_hs(normales_tope)
                res['hs_ext50']    = mins_a_hs(trabajado - base)
    else:  # normal lun-jue
        if jornada_12:
            normales_tope = 9 * 60
            if trabajado <= normales_tope:
                res['hs_normales'] = mins_a_hs(trabajado)
            else:
                res['hs_normales'] = mins_a_hs(normales_tope)
                res['hs_ext50']    = mins_a_hs(trabajado - normales_tope)
        else:
            normales_tope = 9 * 60
            base           = 9 * 60 + 30
            if trabajado <= normales_tope:
                res['hs_normales'] = mins_a_hs(trabajado)
            elif trabajado <= base:
                res['hs_normales'] = mins_a_hs(normales_tope)
            else:
                res['hs_normales'] = mins_a_hs(normales_tope)
                res['hs_ext50']    = mins_a_hs(trabajado - base)
    return res

def construir_mapeo(df_personal, df_reloj):
    emps_reloj = df_reloj[['ID de trabajo','Nombre']].drop_duplicates()
    mapeo = {}
    sin_mapeo = []
    for _, emp in df_personal.iterrows():
        legajo = emp['LEGAJO']
        nombre = emp['NOMBRE']
        if legajo in emps_reloj['ID de trabajo'].values:
            mapeo[legajo] = legajo
            continue
        candidatos = []
        for _, r in emps_reloj.iterrows():
            sim = similitud_nombre(nombre, r['Nombre'])
            if sim >= 0.6:
                candidatos.append((sim, r['ID de trabajo'], r['Nombre']))
        candidatos.sort(reverse=True)
        if not candidatos:
            sin_mapeo.append(legajo)
        else:
            mapeo[legajo] = candidatos[0][1]
    return mapeo, sin_mapeo

def leer_personal(file_bytes):
    df_raw = pd.read_excel(io.BytesIO(file_bytes), header=None)
    hr = None
    for i, row in df_raw.iterrows():
        if 'LEGAJO' in str(row.values):
            hr = i
            break
    if hr is None:
        raise ValueError("No se encontró columna LEGAJO en el archivo de personal")
    df = pd.read_excel(io.BytesIO(file_bytes), header=hr)
    df.columns = ['_', 'N', 'LEGAJO', 'NOMBRE']
    df = df[df['LEGAJO'].apply(lambda x: str(x).replace('.0','').strip().isdigit())].copy()
    df['LEGAJO'] = df['LEGAJO'].astype(int)
    df['NOMBRE'] = df['NOMBRE'].str.strip()
    return df[['LEGAJO','NOMBRE']].reset_index(drop=True)

# ═══════════════════════════════════════════════
#  RUTAS
# ═══════════════════════════════════════════════
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/procesar', methods=['POST'])
def procesar():
    try:
        quincena  = int(request.form['quincena'])
        anio      = int(request.form['anio'])
        mes       = int(request.form['mes'])
        sector    = request.form.get('sector', 'Premoldeado')
        lista_negro_raw = request.form.get('lista_negro', '')
        lista_negro = set()
        for x in lista_negro_raw.split(','):
            x = x.strip()
            if x.isdigit():
                lista_negro.add(int(x))

        if 'reloj' not in request.files or 'personal' not in request.files:
            return jsonify({'error': 'Faltan archivos'}), 400

        reloj_bytes    = request.files['reloj'].read()
        personal_bytes = request.files['personal'].read()

        # Rango de fechas
        if quincena == 1:
            fecha_ini = date(anio, mes, 1)
            fecha_fin = date(anio, mes, 15)
        else:
            fecha_ini = date(anio, mes, 16)
            fecha_fin = (date(anio, mes+1, 1) - timedelta(days=1)) if mes < 12 else date(anio, 12, 31)

        # Leer reloj completo para mapeo, filtrar para cálculo
        df_reloj_full = pd.read_excel(io.BytesIO(reloj_bytes))
        df_reloj_full['fecha'] = pd.to_datetime(df_reloj_full['Fecha de atención']).dt.date
        df_reloj = df_reloj_full[
            (df_reloj_full['fecha'] >= fecha_ini) &
            (df_reloj_full['fecha'] <= fecha_fin)
        ].copy()

        df_personal = leer_personal(personal_bytes)
        mapeo, sin_mapeo = construir_mapeo(df_personal, df_reloj_full)

        # Generar días del período
        dias = []
        d = fecha_ini
        while d <= fecha_fin:
            tipo = clasificar_dia(d)
            if tipo not in ('domingo',):
                dias.append({'fecha': d.isoformat(), 'tipo': tipo,
                             'label': d.strftime('%d/%m'), 'dow': d.weekday()})
            d += timedelta(days=1)

        # Procesar empleados
        empleados = []
        for _, emp in df_personal.iterrows():
            legajo = int(emp['LEGAJO'])
            nombre = emp['NOMBRE']
            id_reloj = mapeo.get(legajo)
            en_negro = legajo in lista_negro
            sin_datos = id_reloj is None

            registros_dia = {}
            if id_reloj is not None:
                regs = df_reloj[df_reloj['ID de trabajo'] == id_reloj]
                for _, reg in regs.iterrows():
                    fecha_str = reg['fecha'].isoformat()
                    registros_dia[fecha_str] = {
                        'entrada': str(reg['Hora de entrada']),
                        'salida':  str(reg['Hora de salida']),
                        'periodo': str(reg['Período']),
                    }

            dias_data = []
            total_norm = total_50 = total_100 = total_enf = total_acc = 0.0

            for dia in dias:
                fecha_str = dia['fecha']
                fecha_obj = date.fromisoformat(fecha_str)
                tipo_dia  = dia['tipo']
                reg = registros_dia.get(fecha_str, {})

                es_descanso = 'Descanso' in reg.get('periodo', '')
                entrada_raw = reg.get('entrada', 'Sin reg.')
                salida_raw  = reg.get('salida',  'Sin reg.')
                entrada_min = parse_hora(entrada_raw)
                salida_min  = parse_hora(salida_raw)

                if es_descanso and entrada_min is None and salida_min is None:
                    dias_data.append({
                        'fecha': fecha_str, 'tipo_dia': tipo_dia,
                        'entrada': '', 'salida': '',
                        'hs_normales': 0, 'hs_ext50': 0, 'hs_ext100': 0,
                        'hs_enfermedad': 0, 'hs_accidente': 0,
                        'estado': 'descanso', 'obs': ''
                    })
                    continue

                calc = calcular_horas_dia(entrada_min, salida_min, tipo_dia)

                if calc['ausente']:
                    estado = 'ausente'
                elif calc['marcacion_incompleta']:
                    estado = 'incompleto'
                else:
                    estado = 'ok'

                dias_data.append({
                    'fecha': fecha_str,
                    'tipo_dia': tipo_dia,
                    'entrada': entrada_raw if entrada_min is not None else '',
                    'salida':  salida_raw  if salida_min  is not None else '',
                    'hs_normales':   calc['hs_normales'],
                    'hs_ext50':      calc['hs_ext50'],
                    'hs_ext100':     calc['hs_ext100'],
                    'hs_enfermedad': 0,
                    'hs_accidente':  0,
                    'estado': estado,
                    'obs': ''
                })
                total_norm += calc['hs_normales']
                total_50   += calc['hs_ext50']
                total_100  += calc['hs_ext100']

            empleados.append({
                'legajo':    legajo,
                'nombre':    nombre,
                'en_negro':  en_negro,
                'sin_datos': sin_datos,
                'jornada_12': False,
                'dias':      dias_data,
                'totales': {
                    'hs_normales':   round(total_norm, 2),
                    'hs_ext50':      round(total_50,   2),
                    'hs_ext100':     round(total_100,  2),
                    'hs_enfermedad': 0,
                    'hs_accidente':  0,
                }
            })

        resultado = {
            'quincena': quincena, 'anio': anio, 'mes': mes,
            'mes_nombre': NOMBRES_MES[mes], 'sector': sector,
            'fecha_ini': fecha_ini.isoformat(),
            'fecha_fin': fecha_fin.isoformat(),
            'dias': dias,
            'empleados': empleados,
            'sin_mapeo': sin_mapeo,
        }

        # Guardar en sesión para descarga posterior
        session['ultimo_resultado'] = json.dumps(resultado, default=str)
        return jsonify(resultado)

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'detalle': traceback.format_exc()}), 500


@app.route('/api/descargar', methods=['POST'])
def descargar():
    try:
        data = request.get_json()
        empleados = data['empleados']
        meta = data['meta']

        anio      = meta['anio']
        mes       = meta['mes']
        quincena  = meta['quincena']
        sector    = meta.get('sector', '')
        fecha_ini = meta['fecha_ini']
        fecha_fin = meta['fecha_fin']
        nombre_mes = NOMBRES_MES[mes]
        q_str = f"Q{quincena}"

        # Colores
        AZUL_OSC  = "1F4E79"
        AZUL_MED  = "2E75B6"
        AZUL_CLAR = "DEEAF1"
        AMARILLO  = "FFE699"
        ROJO_CLAR = "FFD7D7"
        NARANJA   = "FFF2CC"

        def s(cell, bold=False, ctxt="000000", fill=None, sz=9, ha="center", wrap=False):
            cell.font = Font(bold=bold, color=ctxt, name="Arial", size=sz)
            if fill:
                cell.fill = PatternFill("solid", start_color=fill, end_color=fill)
            cell.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
            thin = Side(style="thin", color="B8CCE4")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        wb = Workbook()

        # ─── HOJA 1: Resumen quincena ───
        ws = wb.active
        ws.title = f"Resumen {q_str}"
        titulo = (f"PLANILLA DE HORAS — {sector.upper()}  |  "
                  f"{nombre_mes.upper()} {anio}  |  QUINCENA {quincena}  "
                  f"({fecha_ini[8:10]} al {fecha_fin[8:10]})")
        ws.merge_cells("A1:K1")
        ws["A1"] = titulo
        s(ws["A1"], bold=True, ctxt="FFFFFF", fill=AZUL_OSC, sz=12)
        ws.row_dimensions[1].height = 24

        cols = [
            ("LEGAJO",9), ("APELLIDO Y NOMBRE",32),
            ("HS\nNORMALES",12), ("EXT 50%\nBLANCO",12), ("EXT 100%\nBLANCO",12),
            ("EXT 50%\nNEGRO",12), ("EXT 100%\nNEGRO",12),
            ("HS\nENFERMEDAD",12), ("HS\nACCIDENTE",12),
            ("TOTAL HS\nTRABAJADAS",13), ("OBSERVACIONES",40),
        ]
        for ci,(t,a) in enumerate(cols,1):
            c = ws.cell(row=2, column=ci, value=t)
            s(c, bold=True, ctxt="FFFFFF", fill=AZUL_MED, wrap=True)
            ws.column_dimensions[get_column_letter(ci)].width = a
        ws.row_dimensions[2].height = 30

        tots = {k:0.0 for k in ['norm','e50','e100','e50n','e100n','enf','acc']}

        for ri, emp in enumerate(empleados):
            row = ri + 3
            t = emp['totales']
            en_negro  = emp.get('en_negro', False)
            e50_b  = t['hs_ext50']   if not en_negro else 0.0
            e100_b = t['hs_ext100']  if not en_negro else 0.0
            e50_n  = t['hs_ext50']   if en_negro else 0.0
            e100_n = t['hs_ext100']  if en_negro else 0.0
            total_hs = round(t['hs_normales']+e50_b+e100_b+e50_n+e100_n, 2)

            obs_list = [d.get('obs','') for d in emp['dias'] if d.get('obs','')]
            obs = ' | '.join(obs_list)

            if emp.get('sin_datos'):
                fila_fill = ROJO_CLAR
            elif any(d['estado']=='incompleto' for d in emp['dias']):
                fila_fill = NARANJA
            elif ri % 2 == 0:
                fila_fill = AZUL_CLAR
            else:
                fila_fill = None

            vals = [emp['legajo'], emp['nombre'],
                    t['hs_normales'], e50_b, e100_b, e50_n, e100_n,
                    t['hs_enfermedad'], t['hs_accidente'], total_hs, obs]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=val)
                fill_use = AMARILLO if ci in (6,7) and en_negro and (e50_n>0 or e100_n>0) else fila_fill
                s(c, fill=fill_use, ha="left" if ci in (2,11) else "center")

            tots['norm'] += t['hs_normales']
            tots['e50']  += e50_b
            tots['e100'] += e100_b
            tots['e50n'] += e50_n
            tots['e100n']+= e100_n
            tots['enf']  += t['hs_enfermedad']
            tots['acc']  += t['hs_accidente']

        rt = len(empleados) + 3
        ws.merge_cells(f"A{rt}:B{rt}")
        c = ws.cell(row=rt, column=1, value="TOTALES")
        s(c, bold=True, ctxt="FFFFFF", fill=AZUL_OSC)
        for ci, val in enumerate([tots['norm'],tots['e50'],tots['e100'],
                                   tots['e50n'],tots['e100n'],tots['enf'],tots['acc'],
                                   round(sum(tots.values()),2),''], 3):
            c = ws.cell(row=rt, column=ci, value=round(val,2) if isinstance(val,float) else val)
            s(c, bold=True, ctxt="FFFFFF", fill=AZUL_OSC)
        ws.freeze_panes = "A3"

        # ─── HOJA 2: Para estudio contable ───
        ws2 = wb.create_sheet("Para Estudio Contable")
        ws2.merge_cells("A1:G1")
        ws2["A1"] = f"HORAS EN BLANCO — {sector.upper()}  |  {nombre_mes.upper()} {anio}  |  Q{quincena}"
        s(ws2["A1"], bold=True, ctxt="FFFFFF", fill=AZUL_OSC, sz=12)
        for ci,(t,a) in enumerate([("LEGAJO",9),("APELLIDO Y NOMBRE",32),
                                    ("HS NORMALES",13),("HS EXT 50%",13),
                                    ("HS EXT 100%",13),("HS ENFERMED.",13),("TOTAL HS",13)],1):
            c = ws2.cell(row=2, column=ci, value=t)
            s(c, bold=True, ctxt="FFFFFF", fill=AZUL_MED)
            ws2.column_dimensions[get_column_letter(ci)].width = a
        for ri, emp in enumerate(empleados):
            row = ri+3
            t = emp['totales']
            en_negro = emp.get('en_negro', False)
            e50_b  = t['hs_ext50']  if not en_negro else 0.0
            e100_b = t['hs_ext100'] if not en_negro else 0.0
            total_b = round(t['hs_normales']+e50_b+e100_b, 2)
            fill2 = AZUL_CLAR if ri%2==0 else None
            for ci,val in enumerate([emp['legajo'],emp['nombre'],
                                      t['hs_normales'],e50_b,e100_b,
                                      t['hs_enfermedad'],total_b],1):
                c = ws2.cell(row=row, column=ci, value=val)
                s(c, fill=fill2, ha="left" if ci==2 else "center")
        ws2.freeze_panes = "A3"

        # ─── HOJA 3: Incidencias ───
        ws3 = wb.create_sheet("Incidencias")
        ws3.merge_cells("A1:D1")
        ws3["A1"] = "REGISTRO DE INCIDENCIAS"
        s(ws3["A1"], bold=True, ctxt="FFFFFF", fill=AZUL_OSC, sz=12)
        for ci,(t,a) in enumerate([("LEGAJO",9),("NOMBRE",32),("TIPO",22),("DETALLE",50)],1):
            c = ws3.cell(row=2, column=ci, value=t)
            s(c, bold=True, ctxt="FFFFFF", fill=AZUL_MED)
            ws3.column_dimensions[get_column_letter(ci)].width = a
        inc_row = 3
        for emp in empleados:
            tipo = None
            detalle = ''
            if emp.get('sin_datos'):
                tipo = 'Sin datos en reloj'
            elif any(d['estado']=='incompleto' for d in emp['dias']):
                tipo = 'Marcación incompleta'
                detalle = ' | '.join(d['fecha'] for d in emp['dias'] if d['estado']=='incompleto')
            if tipo:
                fill3 = ROJO_CLAR if 'Sin datos' in tipo else NARANJA
                for ci,val in enumerate([emp['legajo'],emp['nombre'],tipo,detalle],1):
                    c = ws3.cell(row=inc_row, column=ci, value=val)
                    s(c, fill=fill3, ha="left" if ci in (2,4) else "center")
                inc_row += 1
        if inc_row == 3:
            ws3.cell(row=3, column=1, value="Sin incidencias registradas")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"Quincena_Q{quincena}_{nombre_mes}_{anio}_{sector}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'detalle': traceback.format_exc()}), 500


if __name__ == '__main__':
    app.run(debug=True)
